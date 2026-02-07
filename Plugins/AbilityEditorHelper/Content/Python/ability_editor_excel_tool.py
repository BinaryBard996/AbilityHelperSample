# AbilityEditorExcelTools.py
# 工具功能：
# Schema驱动工作流（从 Schema 目录读取 <StructName>.schema.json 生成模板并导出）
# 1) 基于Schema生成Excel/CSV模板（支持嵌套结构、数组、枚举等）
# 2) 基于Schema从Excel/CSV导出为可被FJsonObjectConverter反序列化的Json数组
# 3) 若未安装openpyxl，自动回退为CSV模板/导出

import json
import os
import sys
from typing import Optional

try:
    import openpyxl  # type: ignore
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# Unreal 相关（在编辑器Python中运行）
try:
    import unreal  # type: ignore
    UE_AVAILABLE = True
except Exception:
    UE_AVAILABLE = False

def _ensure_dir(path: str):
    d = os.path.dirname(path)
    if d and not os.path.exists(d):
        os.makedirs(d, exist_ok=True)

# =========================
# Schema 驱动（新增）
# =========================

def _log(msg: str):
    if UE_AVAILABLE:
        unreal.log(msg)
    else:
        print(msg)

def _warn(msg: str):
    if UE_AVAILABLE:
        unreal.log_warning(msg)
    else:
        print(f"[WARN] {msg}")

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    """
    安全地截断 Excel 工作表名称（Excel 限制 31 字符）
    """
    if len(name) <= max_len:
        return name
    return name[:max_len]

def _find_matching_sub_sheet(existing_sheets: list, sub_name: str, matched_main_sheet: str) -> str:
    """
    智能匹配子表名称
    Args:
        existing_sheets: 现有工作表名称列表
        sub_name: 子表后缀名（如 "Modifiers"）
        matched_main_sheet: 匹配到的主表名称
    Returns:
        匹配到的现有子表名称，如果没找到返回空字符串
    """
    # 首先尝试使用匹配到的主表名称查找
    candidate = _safe_sheet_name(f"{matched_main_sheet}.{sub_name}")
    if candidate in existing_sheets:
        return candidate

    # 如果没找到，查找以 ".{sub_name}" 结尾的工作表（处理截断情况）
    suffix = f".{sub_name}"
    truncated_suffix = suffix[:31] if len(suffix) > 31 else suffix
    for sheet in existing_sheets:
        if sheet.endswith(truncated_suffix) or sheet.endswith(suffix[:len(sheet.split(".")[-1]) + 1] if "." in sheet else ""):
            # 检查是否是子表（包含 "."）
            if "." in sheet:
                # 提取子表名称部分
                sheet_sub_name = sheet.split(".")[-1]
                if sub_name.startswith(sheet_sub_name) or sheet_sub_name.startswith(sub_name[:len(sheet_sub_name)]):
                    return sheet
    return ""

def _schema_dir_default() -> str:
    # 插件 Content/Python/Schema
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(here, "Schema")

def _struct_name_from_struct_path(struct_path: str) -> str:
    # /Script/Module.StructName -> StructName
    if not struct_path:
        return ""
    if "." in struct_path:
        return struct_path.split(".")[-1].strip()
    return struct_path.strip()

def _load_schema(schema_name_or_path: str, schema_dir: Optional[str] = None) -> dict:
    if not schema_name_or_path:
        raise ValueError("schema_name_or_path 为空")

    schema_dir = schema_dir or _schema_dir_default()

    if os.path.exists(schema_name_or_path) and schema_name_or_path.lower().endswith(".json"):
        schema_path = schema_name_or_path
    else:
        # 允许传入 StructName（如 GameplayEffectConfig）
        base = schema_name_or_path
        if base.lower().endswith(".schema.json"):
            file_name = base
        else:
            file_name = f"{base}.schema.json"
        schema_path = os.path.join(schema_dir, file_name)

    if not os.path.exists(schema_path):
        raise FileNotFoundError(f"Schema 文件不存在：{schema_path}")

    with open(schema_path, "r", encoding="utf-8") as f:
        return json.load(f)

def _get_special_rule(schema: dict, type_name_or_path: str) -> str:
    """
    schema["specialRules"] 里键是短名（例如 GameplayTagContainer / GameplayAttribute）
    """
    rules = schema.get("specialRules") or {}
    key = type_name_or_path
    if "/" in key and "." in key:
        # /Script/GameplayTags.GameplayTagContainer -> GameplayTagContainer
        key = _struct_name_from_struct_path(key)
    return rules.get(key, "")

def _excel_col_name(field: dict) -> str:
    # 优先 ExcelName，否则用字段名
    n = (field.get("excelName") or "").strip()
    if n:
        return n
    return str(field.get("name") or "").strip()

def _is_primitive_array(field: dict) -> bool:
    """判断是否是基本类型数组（非 struct 数组）"""
    if (field.get("kind") or "").strip() != "array":
        return False
    inner_kind = (field.get("innerKind") or "").strip()
    return inner_kind != "struct"

def _field_hint(schema: dict, field: dict) -> str:
    # 优先 ExcelHint；枚举则提示 enumValues；特殊类型给约定提示
    if field.get("bExcelIgnore"):
        return ""

    excel_hint = (field.get("excelHint") or "").strip()
    if excel_hint:
        return excel_hint

    kind = (field.get("kind") or "").strip()
    if kind == "enum":
        vals = field.get("enumValues") or []
        if vals:
            return " | ".join([str(v) for v in vals])

    if kind == "struct":
        struct_path = field.get("structPath") or ""
        rule = _get_special_rule(schema, struct_path)
        if rule == "tag_container_rule":
            return "用逗号/分号分隔：TagA, TagB.Sub"
        if rule == "attribute_rule":
            return "格式：/Script/Module.Class:Property"

    # 基本类型数组提示
    if kind == "array":
        inner_kind = (field.get("innerKind") or "").strip()
        if inner_kind != "struct":
            inner_enum_vals = field.get("innerEnumValues") or []
            if inner_enum_vals:
                return f"用逗号分隔，可选值：{' | '.join([str(v) for v in inner_enum_vals])}"
            return f"用逗号分隔多个值"

    return ""

def _schema_fields(schema: dict) -> list:
    return schema.get("fields") or []

def _schema_struct_name(schema: dict) -> str:
    return _struct_name_from_struct_path(schema.get("structPath") or "")

def _get_enum_values_for_field(field: dict) -> list:
    """获取字段的枚举值列表（支持enum类型和enum数组）"""
    kind = (field.get("kind") or "").strip()

    if kind == "enum":
        return field.get("enumValues") or []

    # 数组的内部元素是枚举
    if kind == "array":
        inner_kind = (field.get("innerKind") or "").strip()
        if inner_kind == "enum":
            return field.get("innerEnumValues") or []

    return []

def _apply_enum_validation(ws, col_idx: int, enum_values: list, start_row: int = 3, end_row: int = 1000):
    """
    为Excel列应用下拉列表数据验证
    - col_idx: 列索引（1-based）
    - enum_values: 枚举值列表
    - start_row: 起始行（默认3，跳过标题和提示行）
    - end_row: 结束行
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    if not enum_values:
        return

    # Excel 数据验证的 formula1 有 255 字符限制
    # 如果选项太长，截断并警告
    options_str = ",".join(enum_values)
    if len(options_str) > 250:
        _warn(f"[ExcelTools] 枚举选项过多，下拉列表可能被截断")
        # 尝试截断
        truncated = []
        current_len = 0
        for v in enum_values:
            if current_len + len(v) + 1 > 250:
                break
            truncated.append(v)
            current_len += len(v) + 1
        options_str = ",".join(truncated)

    dv = DataValidation(
        type="list",
        formula1=f'"{options_str}"',
        allow_blank=True,
        showDropDown=False,  # False = 显示下拉箭头
        showErrorMessage=True,
        errorTitle="无效输入",
        error="请从下拉列表中选择有效的值"
    )

    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    dv.add(cell_range)
    ws.add_data_validation(dv)

def _validate_enum_value(field: dict, value, row_name: str) -> str:
    """
    验证枚举值是否有效

    Args:
        field: Schema 字段定义
        value: 单元格值
        row_name: 行名（用于错误定位）

    Returns:
        错误信息字符串，如果验证通过则返回空字符串
    """
    kind = (field.get("kind") or "").strip()
    if kind != "enum":
        return ""

    enum_values = field.get("enumValues") or []
    if not enum_values:
        return ""  # 没有定义枚举值列表，跳过验证

    field_name = field.get("name", "Unknown")

    # 获取字符串值
    if value is None or value == "":
        str_value = ""
    else:
        str_value = str(value).strip()

    # 检查空值
    if not str_value:
        return f"[{row_name}] 枚举字段 '{field_name}' 未配置值，可选值: {', '.join(enum_values)}"

    # 检查值有效性
    if str_value not in enum_values:
        return f"[{row_name}] 枚举字段 '{field_name}' 值 '{str_value}' 不在可选值列表中，可选值: {', '.join(enum_values)}"

    return ""

def _read_existing_xlsx_data(xlsx_path: str) -> tuple:
    """
    读取现有Excel文件的数据和列宽
    返回:
        - data_dict: {sheet_name: [{col_name: value, ...}, ...]}
        - col_widths_dict: {sheet_name: {col_name: width, ...}}
    跳过第二行（提示行）
    """
    # 规范化路径
    xlsx_path = os.path.normpath(xlsx_path)
    _log(f"[ExcelTools] _read_existing_xlsx_data: 检查文件 {xlsx_path}")
    if not os.path.exists(xlsx_path):
        _log(f"[ExcelTools] _read_existing_xlsx_data: 文件不存在")
        return {}, {}

    _log(f"[ExcelTools] _read_existing_xlsx_data: 文件存在，开始读取")
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        _log(f"[ExcelTools] _read_existing_xlsx_data: 工作簿加载成功，工作表: {wb.sheetnames}")
        data_result = {}
        col_widths_result = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.rows)
            if not rows:
                data_result[sheet_name] = []
                col_widths_result[sheet_name] = {}
                continue

            # 第一行是表头
            headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]

            # 读取列宽（按列名索引）
            col_widths = {}
            for col_idx, header in enumerate(headers):
                if header:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(col_idx + 1)
                    if col_letter in ws.column_dimensions:
                        width = ws.column_dimensions[col_letter].width
                        if width is not None and width > 0:
                            col_widths[header] = width
            col_widths_result[sheet_name] = col_widths

            # 读取数据
            data = []
            # 从第3行开始读取数据（跳过表头和提示行）
            for row_idx, row in enumerate(rows):
                if row_idx < 2:  # 跳过表头和提示行
                    continue
                # 跳过空行
                if all((c.value is None or str(c.value).strip() == "") for c in row):
                    continue
                row_data = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        row_data[headers[col_idx]] = cell.value
                data.append(row_data)

            data_result[sheet_name] = data

        return data_result, col_widths_result
    except Exception as e:
        _warn(f"[ExcelTools] 读取现有文件失败：{e}，将创建新文件")
        return {}, {}

def _write_xlsx_template_from_schema(schema: dict, out_xlsx_path: str, schema_dir: Optional[str] = None, preserve_data: bool = True):
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("未安装openpyxl，无法生成xlsx。请安装后重试，或使用CSV回退。")

    # 规范化输出路径
    out_xlsx_path = os.path.normpath(out_xlsx_path)
    _log(f"[ExcelTools] 输出路径（规范化后）: {out_xlsx_path}")

    schema_dir = schema_dir or _schema_dir_default()
    struct_name = _schema_struct_name(schema)
    main_sheet_name = struct_name if struct_name else "Main"

    # 读取现有数据和列宽（如果文件存在且需要保留数据）
    existing_data = {}
    existing_col_widths = {}
    if preserve_data:
        _log(f"[ExcelTools] preserve_data=True，尝试读取现有文件：{out_xlsx_path}")
        existing_data, existing_col_widths = _read_existing_xlsx_data(out_xlsx_path)
        if existing_data:
            _log(f"[ExcelTools] 检测到现有文件，将保留已有数据和列宽")
            for sheet_name, rows in existing_data.items():
                _log(f"[ExcelTools]   - 工作表 '{sheet_name}': {len(rows)} 行数据")
                if rows:
                    _log(f"[ExcelTools]     列名: {list(rows[0].keys())}")
        else:
            _log(f"[ExcelTools] 未检测到现有数据（文件不存在或为空）")

    wb = openpyxl.Workbook()
    ws_main = wb.active
    ws_main.title = main_sheet_name

    # 收集有 ExcelSheet 元数据的字段，按 ExcelSheet 分组
    excel_sheet_fields = {}  # ExcelSheet -> [fields...]
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        excel_sheet = (f.get("excelSheet") or "").strip()
        if excel_sheet:
            excel_sheet_fields.setdefault(excel_sheet, []).append(f)

    # 主表：Name + 非数组字段 + 基本类型数组字段（排除有 ExcelSheet 的字段）
    main_fields = [{"name": "Name", "kind": "string"}]  # DataTable行名/主键
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        # 跳过有 ExcelSheet 元数据的字段（会创建单独子表）
        excel_sheet = (f.get("excelSheet") or "").strip()
        if excel_sheet:
            continue
        # 跳过 array<struct>（会创建子表），保留基本类型数组
        if (f.get("kind") or "") == "array" and not _is_primitive_array(f):
            continue
        main_fields.append(f)

    headers = [_excel_col_name(f) for f in main_fields]
    ws_main.append(headers)
    ws_main.append([_field_hint(schema, f) for f in main_fields])

    # 写入现有数据（如果有）
    _log(f"[ExcelTools] 查找主表数据，sheet名称: '{main_sheet_name}'")
    _log(f"[ExcelTools] 现有数据的sheet名称: {list(existing_data.keys())}")

    # 智能匹配主表：如果精确匹配失败，尝试查找现有的主表（不含 "." 的工作表）
    main_existing = existing_data.get(main_sheet_name, [])
    matched_main_sheet = main_sheet_name
    if not main_existing and existing_data:
        # 查找不含 "." 的工作表（主表候选）
        main_sheet_candidates = [s for s in existing_data.keys() if "." not in s]
        if len(main_sheet_candidates) == 1:
            matched_main_sheet = main_sheet_candidates[0]
            main_existing = existing_data.get(matched_main_sheet, [])
            _log(f"[ExcelTools] 使用现有主表 '{matched_main_sheet}' 的数据（智能匹配）")

    if main_existing:
        _log(f"[ExcelTools] 新表头: {headers}")
        _log(f"[ExcelTools] 现有数据第一行的列名: {list(main_existing[0].keys()) if main_existing else []}")
    for row_data in main_existing:
        row_values = []
        for header in headers:
            row_values.append(row_data.get(header, None))
        ws_main.append(row_values)

    if main_existing:
        _log(f"[ExcelTools] 主表保留了 {len(main_existing)} 行数据")
    else:
        _log(f"[ExcelTools] 主表没有找到匹配的现有数据")

    # 为枚举列添加下拉列表验证（动态计算结束行）
    # 至少覆盖现有数据 + 500 行余量，最少1000行
    main_end_row = max(1000, len(main_existing) + 3 + 500)
    for col_idx, field in enumerate(main_fields, start=1):
        enum_values = _get_enum_values_for_field(field)
        if enum_values:
            _apply_enum_validation(ws_main, col_idx, enum_values, end_row=main_end_row)

    # 应用主表列宽（如果有保存的列宽，使用匹配到的主表名称）
    main_col_widths = existing_col_widths.get(matched_main_sheet, {})
    if main_col_widths:
        _log(f"[ExcelTools] 应用主表列宽，来源工作表: '{matched_main_sheet}'")
        from openpyxl.utils import get_column_letter
        for col_idx, header in enumerate(headers, start=1):
            if header in main_col_widths:
                col_letter = get_column_letter(col_idx)
                ws_main.column_dimensions[col_letter].width = main_col_widths[header]

    # ExcelSheet 子表：按 ExcelSheet 元数据分组的字段
    # 子表只使用后缀名，避免 Excel 31 字符限制
    for excel_sheet_name, sheet_fields in excel_sheet_fields.items():
        sub_sheet_name = _safe_sheet_name(excel_sheet_name)
        ws_ext = wb.create_sheet(sub_sheet_name)

        # ParentName 作为外键关联主表
        ext_fields = [{"name": "ParentName", "kind": "string"}]
        ext_fields.extend(sheet_fields)

        ext_headers = [_excel_col_name(x) for x in ext_fields]
        ws_ext.append(ext_headers)
        ws_ext.append([_field_hint(schema, x) for x in ext_fields])

        # 智能匹配子表数据（优先精确匹配，然后尝试旧格式 MainSheet.SubName）
        ext_existing = existing_data.get(sub_sheet_name, [])
        matched_ext_sheet = sub_sheet_name
        if not ext_existing:
            matched_ext_sheet = _find_matching_sub_sheet(list(existing_data.keys()), excel_sheet_name, matched_main_sheet)
            ext_existing = existing_data.get(matched_ext_sheet, []) if matched_ext_sheet else []

        # 写入子表现有数据（如果有）
        for row_data in ext_existing:
            row_values = []
            for header in ext_headers:
                row_values.append(row_data.get(header, None))
            ws_ext.append(row_values)

        if ext_existing:
            _log(f"[ExcelTools] ExcelSheet 子表 {sub_sheet_name} 保留了 {len(ext_existing)} 行数据（来源: '{matched_ext_sheet}'）")

        # 为子表的枚举列添加下拉列表验证
        ext_end_row = max(1000, len(ext_existing) + 3 + 500)
        for col_idx, sf in enumerate(ext_fields, start=1):
            enum_values = _get_enum_values_for_field(sf)
            if enum_values:
                _apply_enum_validation(ws_ext, col_idx, enum_values, end_row=ext_end_row)

        # 应用子表列宽（如果有保存的列宽）
        ext_col_widths = existing_col_widths.get(matched_ext_sheet, {}) if matched_ext_sheet else {}
        if ext_col_widths:
            from openpyxl.utils import get_column_letter
            for col_idx, header in enumerate(ext_headers, start=1):
                if header in ext_col_widths:
                    col_letter = get_column_letter(col_idx)
                    ws_ext.column_dimensions[col_letter].width = ext_col_widths[header]

    # 子表：每个 array<struct> 一个Sheet
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        if (f.get("kind") or "") != "array":
            continue

        # 基本类型数组已在主表处理，跳过
        if _is_primitive_array(f):
            continue

        inner_struct_path = (f.get("innerStructPath") or "").strip()
        inner_struct_name = _struct_name_from_struct_path(inner_struct_path)
        if not inner_struct_name:
            _warn(f"[ExcelTools][Schema] 子表字段缺少 innerStructPath：{f.get('name')}")
            continue

        inner_schema = _load_schema(inner_struct_name, schema_dir=schema_dir)

        # 子表只使用字段名，避免 Excel 31 字符限制
        field_name = str(f.get('name') or '').strip()
        sub_sheet_name = _safe_sheet_name(field_name)
        ws_sub = wb.create_sheet(sub_sheet_name)

        sub_fields = [{"name": "ParentName", "kind": "string"}]
        for sf in _schema_fields(inner_schema):
            if sf.get("bExcelIgnore"):
                continue
            if (sf.get("kind") or "") == "array":
                _warn(f"[ExcelTools][Schema] 子结构体内的数组字段暂不支持：{inner_struct_name}.{sf.get('name')}")
                continue
            sub_fields.append(sf)

        sub_headers = [_excel_col_name(x) for x in sub_fields]
        ws_sub.append(sub_headers)
        ws_sub.append([_field_hint(inner_schema, x) for x in sub_fields])

        # 智能匹配子表数据（优先精确匹配，然后尝试旧格式 MainSheet.SubName）
        sub_existing = existing_data.get(sub_sheet_name, [])
        matched_sub_sheet = sub_sheet_name
        if not sub_existing:
            matched_sub_sheet = _find_matching_sub_sheet(list(existing_data.keys()), field_name, matched_main_sheet)
            sub_existing = existing_data.get(matched_sub_sheet, []) if matched_sub_sheet else []

        # 写入子表现有数据（如果有）
        for row_data in sub_existing:
            row_values = []
            for header in sub_headers:
                row_values.append(row_data.get(header, None))
            ws_sub.append(row_values)

        if sub_existing:
            _log(f"[ExcelTools] 子表 {sub_sheet_name} 保留了 {len(sub_existing)} 行数据（来源: '{matched_sub_sheet}'）")

        # 为子表的枚举列添加下拉列表验证（动态计算结束行）
        sub_end_row = max(1000, len(sub_existing) + 3 + 500)
        for col_idx, sf in enumerate(sub_fields, start=1):
            enum_values = _get_enum_values_for_field(sf)
            if enum_values:
                _apply_enum_validation(ws_sub, col_idx, enum_values, end_row=sub_end_row)

        # 应用子表列宽（如果有保存的列宽）
        sub_col_widths = existing_col_widths.get(matched_sub_sheet, {}) if matched_sub_sheet else {}
        if sub_col_widths:
            from openpyxl.utils import get_column_letter
            for col_idx, header in enumerate(sub_headers, start=1):
                if header in sub_col_widths:
                    col_letter = get_column_letter(col_idx)
                    ws_sub.column_dimensions[col_letter].width = sub_col_widths[header]

    _ensure_dir(out_xlsx_path)
    wb.save(out_xlsx_path)

def _write_csv_template_from_schema(schema: dict, out_dir_or_file: str, schema_dir: Optional[str] = None):
    """
    CSV 回退策略：
      - 主表：<StructName>.csv
      - 子表：<StructName>.<ArrayFieldName>.csv
    """
    import csv

    schema_dir = schema_dir or _schema_dir_default()
    struct_name = _schema_struct_name(schema)
    main_sheet_name = struct_name if struct_name else "Main"

    # 输出目录判定
    out_dir = out_dir_or_file
    if out_dir_or_file.lower().endswith(".xlsx") or out_dir_or_file.lower().endswith(".csv"):
        out_dir = os.path.dirname(out_dir_or_file) or "."

    os.makedirs(out_dir, exist_ok=True)

    # 主表：Name + 非数组字段 + 基本类型数组字段
    main_fields = [{"name": "Name", "kind": "string"}]
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        # 跳过 array<struct>（会创建子表），保留基本类型数组
        if (f.get("kind") or "") == "array" and not _is_primitive_array(f):
            continue
        main_fields.append(f)

    main_csv = os.path.join(out_dir, f"{main_sheet_name}.csv")
    with open(main_csv, "w", newline="", encoding="utf-8") as fp:
        w = csv.writer(fp)
        w.writerow([_excel_col_name(f) for f in main_fields])
        w.writerow([_field_hint(schema, f) for f in main_fields])

    # 子表：每个 array<struct> 一个CSV
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        if (f.get("kind") or "") != "array":
            continue
        # 基本类型数组已在主表处理，跳过
        if _is_primitive_array(f):
            continue

        inner_struct_name = _struct_name_from_struct_path((f.get("innerStructPath") or "").strip())
        if not inner_struct_name:
            continue
        inner_schema = _load_schema(inner_struct_name, schema_dir=schema_dir)

        sub_fields = [{"name": "ParentName", "kind": "string"}]
        for sf in _schema_fields(inner_schema):
            if sf.get("bExcelIgnore"):
                continue
            if (sf.get("kind") or "") == "array":
                continue
            sub_fields.append(sf)

        sub_csv = os.path.join(out_dir, f"{main_sheet_name}.{str(f.get('name') or '').strip()}.csv")
        with open(sub_csv, "w", newline="", encoding="utf-8") as fp:
            w = csv.writer(fp)
            w.writerow([_excel_col_name(x) for x in sub_fields])
            w.writerow([_field_hint(inner_schema, x) for x in sub_fields])

def generate_excel_template_from_schema(schema_name_or_path: str, out_path: str, schema_dir: Optional[str] = None, preserve_data: bool = True):
    """
    新入口：基于 Schema 生成 Excel/CSV 模板
    - schema_name_or_path：
        - 结构体名（如 GameplayEffectConfig），会在 Schema 目录下查找 GameplayEffectConfig.schema.json
        - 或直接传入 Schema 的绝对/相对路径
    - out_path：
        - .xlsx 且 openpyxl 可用：输出 xlsx
        - 否则：输出 CSV 到 out_path 所在目录（或 out_path 本身作为目录）
    - preserve_data：
        - True（默认）：如果目标文件已存在，保留匹配列名的现有数据
        - False：完全覆盖，不保留任何数据
    """
    schema = _load_schema(schema_name_or_path, schema_dir=schema_dir)

    if out_path.lower().endswith(".xlsx") and OPENPYXL_AVAILABLE:
        _write_xlsx_template_from_schema(schema, out_path, schema_dir=schema_dir, preserve_data=preserve_data)
        _log(f"[ExcelTools][Schema] 模板已生成：{out_path}")
    else:
        _write_csv_template_from_schema(schema, out_path, schema_dir=schema_dir)
        _log(f"[ExcelTools][Schema] 未检测到openpyxl或非xlsx输出，已生成CSV模板到目录：{os.path.dirname(out_path) or out_path}")

def _safe_bool(v, default=False):
    if v is None or v == "":
        return default
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "y", "t", "on"):
        return True
    if s in ("0", "false", "no", "n", "f", "off"):
        return False
    return default

def _get_empty_tag_container() -> dict:
    """生成空的 GameplayTagContainer 结构"""
    return {"GameplayTags": [], "ParentTags": []}

def _get_empty_tag_requirements() -> dict:
    """生成空的 TagRequirementsConfig 结构"""
    return {
        "RequireTags": _get_empty_tag_container(),
        "IgnoreTags": _get_empty_tag_container()
    }

def _get_empty_attribute_based_config() -> dict:
    """生成空的 AttributeBasedModifierConfig 结构"""
    return {
        "BackingAttribute": "",
        "AttributeCalculationType": "AttributeMagnitude",
        "Coefficient": 1.0,
        "PreMultiplyAdditiveValue": 0.0,
        "PostMultiplyAdditiveValue": 0.0
    }

def _get_empty_set_by_caller_config() -> dict:
    """生成空的 SetByCallerModifierConfig 结构"""
    return {
        "DataTag": {"TagName": ""},
        "DataName": ""
    }

def _get_empty_gameplay_tag() -> dict:
    """生成空的 GameplayTag 结构"""
    return {"TagName": ""}

def _parse_gameplay_tag(value) -> dict:
    """解析 GameplayTag 字符串为对象"""
    s = _safe_str(value, "").strip()
    if not s:
        return _get_empty_gameplay_tag()
    # 支持直接填 Tag 名称，如 "GameplayCue.Test.Fire"
    return {"TagName": s}

def _to_scalar_from_cell(schema: dict, field: dict, value):
    kind = (field.get("kind") or "").strip()

    if kind == "bool":
        return _safe_bool(value, False)
    if kind == "int":
        return _safe_num(value, 0)
    if kind in ("float", "double"):
        return float(_safe_num(value, 0))
    if kind in ("string", "name", "text"):
        return _safe_str(value, "")
    if kind == "enum":
        return _safe_str(value, "")
    if kind == "struct":
        struct_path = (field.get("structPath") or "").strip()
        struct_name = _struct_name_from_struct_path(struct_path)
        rule = _get_special_rule(schema, struct_path)

        if rule == "tag_container_rule":
            tags = _split_tag_string(value)
            return _to_tag_container_obj(tags) if tags else _get_empty_tag_container()

        if rule == "attribute_rule":
            s = _safe_str(value, "")
            return _parse_attribute_cell(s) if s else {}

        if rule == "tag_requirements_rule":
            s = _safe_str(value, "")
            return _parse_tag_requirements(value) if s else _get_empty_tag_requirements()

        # 特定结构体的默认值
        if struct_name == "AttributeBasedModifierConfig":
            s = _safe_str(value, "")
            if s.startswith("{") and s.endswith("}"):
                try:
                    return json.loads(s)
                except Exception:
                    pass
            return _get_empty_attribute_based_config()

        if struct_name == "SetByCallerModifierConfig":
            s = _safe_str(value, "")
            if s.startswith("{") and s.endswith("}"):
                try:
                    return json.loads(s)
                except Exception:
                    pass
            return _get_empty_set_by_caller_config()

        # GameplayTag（单个 tag）
        if struct_name == "GameplayTag":
            return _parse_gameplay_tag(value)

        # 通用struct：允许用户直接填一个JSON对象字符串
        s = _safe_str(value, "")
        if s.startswith("{") and s.endswith("}"):
            try:
                return json.loads(s)
            except Exception:
                return {}
        return {}
    return value

def _to_primitive_array_from_cell(field: dict, value) -> list:
    """将逗号分隔的单元格值转换为基本类型数组"""
    inner_kind = (field.get("innerKind") or "").strip()

    # 解析逗号分隔的字符串
    s = _safe_str(value, "")
    if not s:
        return []

    items = [x.strip() for x in s.replace(";", ",").split(",") if x.strip()]

    # 根据 innerKind 转换类型
    result = []
    for item in items:
        if inner_kind == "bool":
            result.append(_safe_bool(item, False))
        elif inner_kind == "int":
            result.append(int(_safe_num(item, 0)))
        elif inner_kind in ("float", "double"):
            result.append(float(_safe_num(item, 0)))
        else:
            # string, name, text, enum, softclass, softobject 等都作为字符串
            result.append(item)

    return result

def _sheet_to_dict_list(ws, skip_second_row_hint: bool = True):
    rows = list(ws.rows)
    if not rows:
        return []
    headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
    data = []
    start_idx = 1
    # 智能检测第二行是否为提示行
    # 提示行的特征：第一列（Name 或 ParentName）为空
    if skip_second_row_hint and len(rows) >= 2:
        second_row = rows[1]
        first_cell_value = second_row[0].value if second_row else None
        # 只有当第一列为空时才跳过（这是提示行的特征）
        if first_cell_value is None or str(first_cell_value).strip() == "":
            start_idx = 2
    for r in rows[start_idx:]:
        # 若整行为空则跳过
        if all((c.value is None or str(c.value).strip() == "") for c in r):
            continue
        row_map = {}
        for i, cell in enumerate(r):
            key = headers[i] if i < len(headers) else f"COL_{i}"
            row_map[key] = cell.value
        data.append(row_map)
    return data

def _read_xlsx_sheet_map(xlsx_path: str) -> dict:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("未安装openpyxl，无法读取xlsx。请安装后重试，或使用CSV回退。")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        out[name] = _sheet_to_dict_list(ws, skip_second_row_hint=True)
    return out

def _read_csv_table(csv_path: str) -> list:
    import csv
    rows = []
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = None
        for i, line in enumerate(reader):
            if i == 0:
                headers = [x.strip() for x in line]
                continue
            # 跳过第二行枚举提示
            if i == 1:
                continue
            if not any((x or "").strip() for x in line):
                continue
            row = {}
            for idx, v in enumerate(line):
                key = headers[idx] if idx < len(headers) else f"COL_{idx}"
                row[key] = v
            rows.append(row)
    return rows

def _read_csv_sheet_map(base_dir_or_csv: str, main_sheet_name: str, array_field_names: list) -> dict:
    base_dir = base_dir_or_csv
    if base_dir_or_csv.lower().endswith(".csv"):
        base_dir = os.path.dirname(base_dir_or_csv) or "."

    out = {}

    main_csv = os.path.join(base_dir, f"{main_sheet_name}.csv")
    if not os.path.exists(main_csv):
        raise FileNotFoundError(f"主表CSV不存在：{main_csv}")
    out[main_sheet_name] = _read_csv_table(main_csv)

    for af in array_field_names:
        sub_name = f"{main_sheet_name}.{af}"
        sub_csv = os.path.join(base_dir, f"{sub_name}.csv")
        if os.path.exists(sub_csv):
            out[sub_name] = _read_csv_table(sub_csv)
        else:
            out[sub_name] = []

    return out

def export_excel_to_json_using_schema(in_path: str, out_json_path: str, schema_name_or_path: str, schema_dir: Optional[str] = None):
    """
    新入口：用 Schema 驱动导出
    - in_path：
        - xlsx 文件
        - 或 CSV 目录/任一csv文件路径
    - out_json_path：输出 json（对象数组）
    - schema_name_or_path：同 generate_excel_template_from_schema
    """
    schema_dir = schema_dir or _schema_dir_default()
    schema = _load_schema(schema_name_or_path, schema_dir=schema_dir)

    # 枚举值校验错误收集
    enum_validation_errors = []

    struct_name = _schema_struct_name(schema)
    main_sheet_name = struct_name if struct_name else "Main"

    # 收集有 ExcelSheet 元数据的字段，按 ExcelSheet 分组
    excel_sheet_fields = {}  # ExcelSheet -> [fields...]
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        excel_sheet = (f.get("excelSheet") or "").strip()
        if excel_sheet:
            excel_sheet_fields.setdefault(excel_sheet, []).append(f)

    # 收集 array<struct> 字段
    array_fields = []
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        if (f.get("kind") or "") != "array":
            continue
        if (f.get("innerKind") or "").strip() != "struct":
            continue
        array_fields.append(f)

    if in_path.lower().endswith(".xlsx"):
        sheet_map = _read_xlsx_sheet_map(in_path)
    else:
        sheet_map = _read_csv_sheet_map(
            in_path,
            main_sheet_name=main_sheet_name,
            array_field_names=[str(f.get("name") or "").strip() for f in array_fields],
        )

    main_rows = sheet_map.get(main_sheet_name) or []

    # ExcelSheet 子表按 ParentName 聚合
    excel_sheet_rows_map = {}  # ExcelSheet -> { ParentName -> row_data }
    for excel_sheet_name in excel_sheet_fields.keys():
        # 优先使用新格式（只有字段名），如果找不到再尝试旧格式（MainSheet.SubName）
        sub_sheet = _safe_sheet_name(excel_sheet_name)
        sub_rows = sheet_map.get(sub_sheet) or []
        if not sub_rows:
            sub_sheet_old = _safe_sheet_name(f"{main_sheet_name}.{excel_sheet_name}")
            sub_rows = sheet_map.get(sub_sheet_old) or []
        by_parent = {}
        for r in sub_rows:
            p = _safe_str(r.get("ParentName"), "")
            if not p:
                continue
            # ExcelSheet 子表每个 ParentName 只有一行（不是数组）
            by_parent[p] = r
        excel_sheet_rows_map[excel_sheet_name] = by_parent

    # 子表按 ParentName 聚合（array<struct>）
    child_rows_map = {}  # arrayFieldName -> { ParentName -> [rows...] }
    for af in array_fields:
        af_name = str(af.get("name") or "").strip()
        # 优先使用新格式（只有字段名），如果找不到再尝试旧格式（MainSheet.SubName）
        sub_sheet = _safe_sheet_name(af_name)
        sub_rows = sheet_map.get(sub_sheet) or []
        if not sub_rows:
            sub_sheet_old = _safe_sheet_name(f"{main_sheet_name}.{af_name}")
            sub_rows = sheet_map.get(sub_sheet_old) or []
        by_parent = {}
        for r in sub_rows:
            p = _safe_str(r.get("ParentName"), "")
            if not p:
                continue
            by_parent.setdefault(p, []).append(r)
        child_rows_map[af_name] = by_parent

    # 主表字段：Name + 非数组字段 + 基本类型数组字段（排除有 ExcelSheet 的字段）
    main_fields = [{"name": "Name", "kind": "string"}]
    for f in _schema_fields(schema):
        if f.get("bExcelIgnore"):
            continue
        # 跳过有 ExcelSheet 元数据的字段（从子表读取）
        excel_sheet = (f.get("excelSheet") or "").strip()
        if excel_sheet:
            continue
        # 跳过 array<struct>（从子表读取），保留基本类型数组
        if (f.get("kind") or "") == "array" and not _is_primitive_array(f):
            continue
        main_fields.append(f)

    result = []
    for r in main_rows:
        name = _safe_str(r.get("Name"), "")
        if not name or name == "Name":
            continue

        item = {"Name": name}

        # 写入非数组字段和基本类型数组字段
        for f in main_fields:
            fn = str(f.get("name") or "").strip()
            if fn == "Name":
                continue
            col = _excel_col_name(f)
            cell_val = r.get(col)
            # 枚举值校验
            enum_error = _validate_enum_value(f, cell_val, name)
            if enum_error:
                enum_validation_errors.append(enum_error)
            # 基本类型数组使用专门的解析函数
            if _is_primitive_array(f):
                arr_val = _to_primitive_array_from_cell(f, cell_val)
                # 空数组也输出，让UE知道这是个数组字段
                item[fn] = arr_val
            else:
                item[fn] = _to_scalar_from_cell(schema, f, cell_val)

        # 写入 ExcelSheet 子表字段
        for excel_sheet_name, sheet_fields in excel_sheet_fields.items():
            ext_row = (excel_sheet_rows_map.get(excel_sheet_name) or {}).get(name)
            for sf in sheet_fields:
                sf_name = str(sf.get("name") or "").strip()
                col = _excel_col_name(sf)
                cell_val = ext_row.get(col) if ext_row else None
                # 枚举值校验
                if ext_row:
                    enum_error = _validate_enum_value(sf, cell_val, f"{name}.{excel_sheet_name}")
                    if enum_error:
                        enum_validation_errors.append(enum_error)
                if _is_primitive_array(sf):
                    item[sf_name] = _to_primitive_array_from_cell(sf, cell_val)
                else:
                    item[sf_name] = _to_scalar_from_cell(schema, sf, cell_val)

        # 写入 array<struct> 子表
        for af in array_fields:
            af_name = str(af.get("name") or "").strip()
            inner_struct_name = _struct_name_from_struct_path((af.get("innerStructPath") or "").strip())
            if not inner_struct_name:
                item[af_name] = []
                continue

            inner_schema = _load_schema(inner_struct_name, schema_dir=schema_dir)
            rows_for_this = (child_rows_map.get(af_name) or {}).get(name) or []
            arr = []
            child_row_idx = 0
            for cr in rows_for_this:
                child_row_idx += 1
                child_obj = {}
                for sf in _schema_fields(inner_schema):
                    if sf.get("bExcelIgnore"):
                        continue
                    if (sf.get("kind") or "") == "array":
                        continue
                    sf_name = str(sf.get("name") or "").strip()
                    col = _excel_col_name(sf)
                    cell_val = cr.get(col)
                    # 枚举值校验
                    enum_error = _validate_enum_value(sf, cell_val, f"{name}.{af_name}[{child_row_idx}]")
                    if enum_error:
                        enum_validation_errors.append(enum_error)
                    child_obj[sf_name] = _to_scalar_from_cell(inner_schema, sf, cell_val)
                arr.append(child_obj)
            item[af_name] = arr

        result.append(item)

    _ensure_dir(out_json_path)
    with open(out_json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=4)

    # 枚举验证结果汇总
    if enum_validation_errors:
        _warn(f"[ExcelTools] 发现 {len(enum_validation_errors)} 个枚举字段问题:")
        for err in enum_validation_errors:
            _warn(f"  - {err}")

    _log(f"[ExcelTools][Schema] 导出完成：{out_json_path}（共{len(result)}条）")

def _split_tag_string(s) -> list:
    if not s:
        return []
    if isinstance(s, str):
        raw = s
    else:
        raw = str(s)
    tags = [t.strip() for t in raw.replace(";", ",").split(",") if t.strip()]
    # 去重保持顺序
    seen = set()
    result = []
    for t in tags:
        if t not in seen:
            seen.add(t)
            result.append(t)
    return result

def _compute_parent_tags(tag: str) -> list:
    # GameplayCue.Test.A -> ["GameplayCue", "GameplayCue.Test"]
    parts = tag.split(".")
    if len(parts) <= 1:
        return []
    parents = []
    for i in range(1, len(parts)):
        parents.append(".".join(parts[:i]))
    return parents

def _to_tag_container_obj(tag_list: list) -> dict:
    gameplay_tags = [{"TagName": t} for t in tag_list]
    parent_set = set()
    for t in tag_list:
        for p in _compute_parent_tags(t):
            parent_set.add(p)
    parent_tags = [{"TagName": p} for p in sorted(parent_set)]
    return {"GameplayTags": gameplay_tags, "ParentTags": parent_tags}

def _parse_attribute_cell(cell_value: str) -> dict:
    """
    期望格式：
      - 完整：/Script/GameplayAbilities.AbilitySystemComponent:OutgoingDuration
      - 或 /Script/Module.Class:Property
    输出：
      {
        "AttributeName": "OutgoingDuration",
        "Attribute": "/Script/GameplayAbilities.AbilitySystemComponent:OutgoingDuration",
        "AttributeOwner": "/Script/CoreUObject.Class'/Script/GameplayAbilities.AbilitySystemComponent'"
      }
    """
    if not cell_value:
        return {}
    s = str(cell_value).strip()
    if ":" not in s:
        raise ValueError(f"Attribute列需使用 /Script/Module.Class:Property 格式，当前：{s}")
    class_path, prop_name = s.split(":", 1)
    class_path = class_path.strip()
    prop_name = prop_name.strip()
    # AttributeOwner 形如：/Script/CoreUObject.Class'/Script/GameplayAbilities.AbilitySystemComponent'
    owner_obj = f"/Script/CoreUObject.Class'{class_path}'"
    return {
        "AttributeName": prop_name,
        "Attribute": f"{class_path}:{prop_name}",
        "AttributeOwner": owner_obj
    }

def _parse_tag_requirements(cell_value) -> dict:
    """
    解析 Tag Requirements 的行内格式
    格式：  "Require:Tag.A,Tag.B|Ignore:Tag.C,Tag.D"
    输出：
      {
        "RequireTags": {"GameplayTags": [...], "ParentTags": [...]},
        "IgnoreTags": {"GameplayTags": [...], "ParentTags": [...]}
      }
    """
    s = _safe_str(cell_value, "")
    if not s:
        return {
            "RequireTags": _to_tag_container_obj([]),
            "IgnoreTags": _to_tag_container_obj([])
        }

    require = []
    ignore = []

    if "|" in s:
        parts = s.split("|")
        for part in parts:
            part = part.strip()
            if part.startswith("Require:"):
                require = _split_tag_string(part[8:])
            elif part.startswith("Ignore:"):
                ignore = _split_tag_string(part[7:])
    else:
        # 没有 | 分隔符时，检查是否有 Require: 或 Ignore: 前缀
        if s.startswith("Require:"):
            require = _split_tag_string(s[8:])
        elif s.startswith("Ignore:"):
            ignore = _split_tag_string(s[7:])
        else:
            # 回退：将整个字符串视为 RequireTags
            require = _split_tag_string(s)

    return {
        "RequireTags": _to_tag_container_obj(require),
        "IgnoreTags": _to_tag_container_obj(ignore)
    }

def _parse_asset_path(cell_value) -> str:
    """
    规范化 UE 资产路径
    接受：
      - 完整路径：/Game/Effects/GE_Base.GE_Base
      - 包路径：/Game/Effects/GE_Base
      - 相对路径：Effects/GE_Base
    返回：规范化的资产路径字符串
    """
    s = _safe_str(cell_value, "")
    if not s:
        return ""

    # 如果不以 / 开头，规范化为 /Game/
    if not s.startswith("/"):
        s = "/Game/" + s

    # 如果只是包路径（没有 .），添加资产名称
    if "." not in s:
        asset_name = s.split("/")[-1]
        s = f"{s}.{asset_name}"

    return s

def _safe_num(v, default=0):
    if v is None or v == "":
        return default
    try:
        if isinstance(v, (int, float)):
            return v
        s = str(v).strip()
        if "." in s:
            return float(s)
        return int(s)
    except Exception:
        return default

def _safe_str(v, default=""):
    if v is None:
        return default
    return str(v).strip()

if __name__ == "__main__":
    # 命令行用法（在外部Python环境下也可调试）
    # 1) Schema 模板：python ability_editor_excel_tool.py schema_template <schema> <输出路径/目录> [--schema-dir DIR]
    # 2) Schema 导出：python ability_editor_excel_tool.py schema_export <schema> <输入xlsx|csv> <输出json> [--schema-dir DIR]
    import argparse
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="cmd")

    p0 = sub.add_parser("schema_template")
    p0.add_argument("schema", help="Schema名称(如 GameplayEffectConfig) 或 .schema.json 路径")
    p0.add_argument("out", help="输出.xlsx路径（无openpyxl时输出CSV到目录）")
    p0.add_argument("--schema-dir", default="", help="Schema目录（默认插件Content/Python/Schema）")
    p0.add_argument("--no-preserve", action="store_true", help="不保留现有数据，完全覆盖")

    p00 = sub.add_parser("schema_export")
    p00.add_argument("schema", help="Schema名称(如 GameplayEffectConfig) 或 .schema.json 路径")
    p00.add_argument("inp", help=".xlsx文件路径 或 CSV目录/任一csv路径")
    p00.add_argument("out", help="输出.json路径")
    p00.add_argument("--schema-dir", default="", help="Schema目录（默认插件Content/Python/Schema）")

    args = parser.parse_args()
    if args.cmd == "schema_template":
        sd = args.schema_dir.strip() or None
        preserve = not getattr(args, 'no_preserve', False)
        generate_excel_template_from_schema(args.schema, args.out, schema_dir=sd, preserve_data=preserve)
    elif args.cmd == "schema_export":
        sd = args.schema_dir.strip() or None
        export_excel_to_json_using_schema(args.inp, args.out, args.schema, schema_dir=sd)
    else:
        print("用法：\n"
              "  schema_template <schema_name|schema_path> <out.xlsx|out_dir> [--schema-dir <dir>] [--no-preserve]\n"
              "  schema_export   <schema_name|schema_path> <in.xlsx|csv_dir> <out.json> [--schema-dir <dir>]")
